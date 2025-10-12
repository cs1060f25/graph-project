#!/usr/bin/env python3
"""Generate papers.xlsx with sample papers from different ArXiv categories."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Sample papers data: 20 papers, 4 from each of 5 categories
papers_data = [
    # Machine Learning (4 papers)
    {
        "Title": "Attention Is All You Need",
        "Authors": "Vaswani, A., Shazeer, N., Parmar, N., et al.",
        "Year": 2017,
        "Folder": "Machine Learning",
        "Link": "https://arxiv.org/abs/1706.03762"
    },
    {
        "Title": "Deep Residual Learning for Image Recognition",
        "Authors": "He, K., Zhang, X., Ren, S., et al.",
        "Year": 2015,
        "Folder": "Machine Learning",
        "Link": "https://arxiv.org/abs/1512.03385"
    },
    {
        "Title": "Neural Machine Translation by Jointly Learning to Align and Translate",
        "Authors": "Bahdanau, D., Cho, K., Bengio, Y.",
        "Year": 2014,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1409.0473"
    },
    {
        "Title": "Dropout: A Simple Way to Prevent Neural Networks from Overfitting",
        "Authors": "Hinton, G. E., Srivastava, N., Krizhevsky, A., et al.",
        "Year": 2012,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1207.0580"
    },
    
    # Computer Vision (4 papers)
    {
        "Title": "You Only Look Once: Unified, Real-Time Object Detection",
        "Authors": "Redmon, J., Divvala, S., Girshick, R., et al.",
        "Year": 2015,
        "Folder": "Computer Vision",
        "Link": "https://arxiv.org/abs/1506.02640"
    },
    {
        "Title": "Mask R-CNN",
        "Authors": "He, K., Gkioxari, G., Dollár, P., et al.",
        "Year": 2017,
        "Folder": "Computer Vision",
        "Link": "https://arxiv.org/abs/1703.06870"
    },
    {
        "Title": "Very Deep Convolutional Networks for Large-Scale Image Recognition",
        "Authors": "Simonyan, K., Zisserman, A.",
        "Year": 2014,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1409.1556"
    },
    {
        "Title": "ImageNet-21K Pretraining for the Masses",
        "Authors": "Ridnik, T., Lawen, H., Noy, A., et al.",
        "Year": 2021,
        "Folder": "",
        "Link": "https://arxiv.org/abs/2106.05969"
    },
    
    # Natural Language Processing (4 papers)
    {
        "Title": "BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding",
        "Authors": "Devlin, J., Chang, M. W., Lee, K., et al.",
        "Year": 2018,
        "Folder": "NLP",
        "Link": "https://arxiv.org/abs/1810.04805"
    },
    {
        "Title": "Language Models are Unsupervised Multitask Learners",
        "Authors": "Radford, A., Wu, J., Child, R., et al.",
        "Year": 2019,
        "Folder": "NLP",
        "Link": "https://arxiv.org/abs/1909.07860"
    },
    {
        "Title": "Sequence to Sequence Learning with Neural Networks",
        "Authors": "Ilya Sutskever, Oriol Vinyals, Quoc V. Le",
        "Year": 2014,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1409.3215"
    },
    {
        "Title": "Transformers: State-of-the-art Natural Language Processing",
        "Authors": "Wolf, T., Debut, L., Sanh, V., et al.",
        "Year": 2019,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1910.03771"
    },
    
    # Computer Vision - Advanced (4 papers)
    {
        "Title": "An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale",
        "Authors": "Dosovitskiy, A., Beyer, L., Kolesnikov, A., et al.",
        "Year": 2020,
        "Folder": "Computer Vision",
        "Link": "https://arxiv.org/abs/2010.11929"
    },
    {
        "Title": "Contrastive Learning with Hard Negative Samples",
        "Authors": "Robinson, J., Chuang, C. Y., Sra, S., et al.",
        "Year": 2020,
        "Folder": "",
        "Link": "https://arxiv.org/abs/2010.04592"
    },
    {
        "Title": "Swin Transformer: Hierarchical Vision Transformer using Shifted Windows",
        "Authors": "Liu, Z., Lin, Y., Cao, Y., et al.",
        "Year": 2021,
        "Folder": "",
        "Link": "https://arxiv.org/abs/2103.14030"
    },
    {
        "Title": "CLIP: Learning Transferable Models for Unsupervised Domain Adaptation",
        "Authors": "Radford, A., Kim, J. W., Hallacy, C., et al.",
        "Year": 2021,
        "Folder": "",
        "Link": "https://arxiv.org/abs/2103.14030"
    },
    
    # Reinforcement Learning (4 papers)
    {
        "Title": "Playing Atari with Deep Reinforcement Learning",
        "Authors": "Mnih, V., Kavukcuoglu, K., Silver, D., et al.",
        "Year": 2013,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1312.5602"
    },
    {
        "Title": "Mastering the game of Go with deep neural networks and tree search",
        "Authors": "Silver, D., Huang, A., Maddison, C. J., et al.",
        "Year": 2016,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1602.01783"
    },
    {
        "Title": "Proximal Policy Optimization Algorithms",
        "Authors": "Schulman, J., Wolski, F., Dhariwal, P., et al.",
        "Year": 2017,
        "Folder": "",
        "Link": "https://arxiv.org/abs/1707.06347"
    },
    {
        "Title": "Deep Reinforcement Learning for Trading",
        "Authors": "Théate, T., Cathcart, D., Ernst, D.",
        "Year": 2020,
        "Folder": "",
        "Link": "https://arxiv.org/abs/2005.12586"
    }
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Papers"

# Create header row
headers = ["Title", "Authors", "Year", "Folder", "Link"]
ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Add data rows
for paper in papers_data:
    ws.append([
        paper["Title"],
        paper["Authors"],
        paper["Year"],
        paper["Folder"],
        paper["Link"]
    ])

# Adjust column widths
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 30

# Set row height for header
ws.row_dimensions[1].height = 25

# Wrap text for all cells
for row in ws.iter_rows(min_row=1, max_row=len(papers_data) + 1):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Save file
output_path = Path(__file__).parent / "public" / "papers.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)

print(f"✅ Generated papers.xlsx with {len(papers_data)} papers")
print(f"📁 File saved to: {output_path}")